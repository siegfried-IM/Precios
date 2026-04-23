"""
Generar Comparativa de precios Cardio Siegfried vs Competidores (Abr 2026).

Entrada:
  - Sin título - Tabla - 21 de abril de 2026.xlsx  (precios AR, 20661 filas)
  - AR_PM_FV_Standard_Apr-21-2026.xlsx             (IQVIA: ATC + MAT)

Salida:
  - Comparativa de precios Abr 2026.xlsx (una hoja por molécula+dosis Siegfried)
"""
from __future__ import annotations

import argparse
import re
import sys
import unicodedata
from pathlib import Path

import pandas as pd
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

sys.stdout.reconfigure(encoding="utf-8")

BASE = Path(r"C:\Users\camarinaro\Downloads\precios cardio")
PRECIOS_XLSX = BASE / "Sin título - Tabla - 21 de abril de 2026.xlsx"
IQVIA_XLSX = BASE / "AR_PM_FV_Standard_Apr-21-2026.xlsx"
OUT_XLSX = BASE / "Comparativa de precios Abr 2026.xlsx"

COL_PVP_ANT = "PVP al 28/02/2026"
COL_PVP_ACT = "PVP al 21/04/2026"
LABEL_PVP_ANT = "PVP Feb 2026"
LABEL_PVP_ACT = "PVP Abr 2026"

# Portfolio cardio Siegfried (primer token del Producto)
# Incluye marcas con ATC C*, B01*, A10* (antidiabéticos cardio-metabólicos)
CARDIO_BRANDS_SIEGFRIED = {
    "DAURAN",    # dapagliflozina (A10BK01)
    "DILATREND", # carvedilol (+ AP, + D con hctz)
    "DIOVAN",    # valsartán (+ D con hctz, + IC)
    "EMPAX",     # empagliflozina (+ MET con metformina)
    "ENTRESTO",  # sacubitrilo+valsartán
    "EXFORGE",   # valsartán+amlodipina (+ D)
    "NEBILET",   # nebivolol (+ D con hctz)
    "PIXABAN",   # apixaban
    "ROXOLAN",   # rosuvastatina (+ PLUS con ezetimibe)
    "SILTRAN",   # sitagliptina (+ MET con metformina)
    "SINTROM",   # acenocumarol
    "TERLOC",    # amlodipina (+ DUO con benazepril)
}

# --- Estilos (replica del template) ---
HDR_FILL = PatternFill("solid", fgColor="1F4E78")  # azul oscuro
HDR_FONT = Font(bold=True, color="FFFFFF")
SIE_FILL = PatternFill("solid", fgColor="FFE699")  # amarillo claro
SIE_FONT = Font(bold=True)
LEADER_FILL = PatternFill("solid", fgColor="C6EFCE")  # verde claro para celda "Líder"

# Anchos de columna — ahora A-M (se agrega col C "Dosis")
COL_WIDTHS = {
    "A": 22.83, "B": 20.83, "C": 12.00, "D": 28.83, "E": 14.83, "F": 14.83,
    "G": 12.83, "H": 8.83, "I": 10.83, "J": 10.83, "K": 12.83,
    "L": 8.83, "M": 6.83,
}

# ---------- Helpers de normalización ----------

def norm_text(s: object) -> str:
    """Uppercase, sin acentos, espacios colapsados."""
    if s is None or (isinstance(s, float) and pd.isna(s)):
        return ""
    s = str(s).strip()
    s = unicodedata.normalize("NFD", s)
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    s = re.sub(r"\s+", " ", s).upper()
    return s


def norm_prod(s: object) -> str:
    """Normalización específica de producto: como norm_text + strip de dosis colgada al final.
    Ej: 'ROSUVAST EZ 10/10' → 'ROSUVAST EZ', 'CARVEDIL 3.125' → 'CARVEDIL'."""
    n = norm_text(s)
    # Strip trailing dose pattern: espacio + número (con decimales/slashes opcionales)
    n = re.sub(r"\s+\d+(?:\.\d+)?(?:/\d+(?:\.\d+)?)*\s*$", "", n).strip()
    return n


def norm_lab(s: object) -> str:
    """Normaliza nombre de laboratorio para matcheo precios↔IQVIA."""
    n = norm_text(s)
    return n


def labs_match_fuzzy(lab_a: str, lab_b: str) -> bool:
    """Compara dos lab norms con tolerancia a truncación típica de IQVIA.
    Ej: 'BOEHRINGER INGELHEIM' ≈ 'BOEHRINGER ING.' → True.
    Regla: primer token debe coincidir; tokens siguientes deben ser mutuamente prefijos (o iguales)."""
    if lab_a == lab_b:
        return True
    a = lab_a.split()
    b = lab_b.split()
    if not a or not b or a[0] != b[0]:
        return False
    n = min(len(a), len(b))
    for i in range(1, n):
        ta = a[i].rstrip(".")
        tb = b[i].rstrip(".")
        if not (ta.startswith(tb) or tb.startswith(ta)):
            return False
    return True


def find_iqvia_lab(precios_lab_norm: str, iqvia_labs: set[str]) -> str | None:
    """Devuelve el lab normalizado tal como aparece en IQVIA."""
    if precios_lab_norm in iqvia_labs:
        return precios_lab_norm
    for il in iqvia_labs:
        if labs_match_fuzzy(precios_lab_norm, il):
            return il
    return None


def extract_mg(presentacion: str) -> str | None:
    """Extrae la dosis principal en mg de la presentación.
    Ej: '10 mg comp.x 30' → '10', '5/10 mg comp...' → '5/10'.
    Ignora unidades internacionales (UI/U.I.) y microgramos no convertidos."""
    if not presentacion:
        return None
    p = str(presentacion).lower().strip()
    # Si viene en UI (heparinas, hormonas), no convertir a mg → skip
    if re.search(r"\bu\.?i\.?\b|\buni(dad|t)", p):
        return None
    # Combinación tipo 5/10, 160/12.5, 25/12.5 (con o sin unidad explícita)
    m = re.search(r"(\d+(?:\.\d+)?(?:/\d+(?:\.\d+)?)+)\s*(?:mg|comp|c[aá]ps|gr|g\b)", p)
    if m:
        return m.group(1)
    # Sólo un valor + mg
    m = re.search(r"(\d+(?:\.\d+)?)\s*mg\b", p)
    if m:
        return m.group(1)
    return None


def extract_count(presentacion: str) -> int | None:
    """Extrae la cantidad por envase. Ej: 'comp.x 30' → 30, 'cáps.x 28' → 28."""
    if not presentacion:
        return None
    m = re.search(r"x\s*(\d+)", str(presentacion).lower())
    return int(m.group(1)) if m else None


def mg_from_iqvia_pack(pack: str) -> str | None:
    """Extrae mg del Pack IQVIA. Maneja dobles y triples combinaciones.
    Formatos:
      - 'DILATREND TABL RAN 5.00MG x 28' → '5'
      - 'SIL MET TA REC 50.00MG/ 500.00MG x 30' → '50/500' (dos MG explícitos)
      - 'DIOVAN-D TABL 12.5MG x 28 /160' → '12.5/160' (dosis 2 al final como /N)
      - 'ENTRESTO TA REV 24MG/ 26mg x 30' → '24/26'
      - 'EMPAX MET TAB REC 1000MG x 60 /5' → '1000/5'
      - 'EXFORGE D TA.160/12.5/ 5.00MG x 28' → '160/12.5/5' (triple combinación)"""
    if not pack:
        return None
    p = str(pack).upper()

    def _fmt(n: float) -> str:
        return str(int(n)) if n.is_integer() else str(n)

    tokens: list[float] = []

    # 1) Patrón multi-dosis: "A/B/C MG" (dos o más números con '/' antes del MG).
    #    Cubre EXFORGE D (triple) y SIL MET (doble con '/' entre ambos).
    multi = re.search(r"((?:\d+(?:\.\d+)?)(?:\s*/\s*\d+(?:\.\d+)?)+)\s*MG", p)
    if multi:
        tokens.extend(float(x) for x in re.findall(r"\d+(?:\.\d+)?", multi.group(1)))
    else:
        # 2) Todos los 'NMG' sueltos (ENTRESTO '24MG/ 26MG' donde ambos tienen MG explícito)
        for m in re.finditer(r"(\d+(?:\.\d+)?)\s*MG", p):
            tokens.append(float(m.group(1)))

    # 3) Sufijo '/N' al final (segunda/tercera dosis implícita, común en DIOVAN-D, EMPAX MET, NEBILET D).
    #    IQVIA a veces deja '/12.' con punto suelto, por eso el '.' es opcional.
    tail = re.search(r"/\s*(\d+(?:\.\d+)?)\.?\s*$", p)
    if tail:
        v = float(tail.group(1))
        if v not in tokens:
            tokens.append(v)
    else:
        # 3b) ROXOLAN PLUS: después del 'x N' hay otra dosis separada por espacios sin '/'
        #     Ej: 'ROXOLAN PLUS TABL RECUBIE 20mg x 30  10' → segunda dosis = 10
        tail2 = re.search(r"X\s*\d+\s{2,}(\d+(?:\.\d+)?)\s*$", p)
        if tail2:
            v = float(tail2.group(1))
            if v not in tokens:
                tokens.append(v)

    if tokens:
        return "/".join(_fmt(n) for n in tokens)

    # 4) Gramos → mg
    m = re.search(r"(\d+(?:\.\d+)?)\s*G\b", p)
    if m:
        return str(float(m.group(1)) * 1000)
    return None


def mg_variants(mg: str | None) -> list[str]:
    """Devuelve variantes equivalentes de una cadena de mg para matching flexible.
    '24/26' → ['24/26', '26/24', '50']
    '160/12.5' → ['160/12.5', '12.5/160', '172.5']
    '5' → ['5']
    None → []"""
    if not mg:
        return []
    variants = [mg]
    if "/" in mg:
        parts = mg.split("/")
        # Orden invertido
        variants.append("/".join(reversed(parts)))
        # Suma (para combinaciones que en precios aparecen sumadas, ej ENTRESTO)
        try:
            s = sum(float(p) for p in parts)
            variants.append(str(int(s)) if s.is_integer() else str(s))
        except ValueError:
            pass
    return variants


def count_from_iqvia_pack(pack: str) -> int | None:
    if not pack:
        return None
    m = re.search(r"X\s*(\d+)", str(pack).upper())
    return int(m.group(1)) if m else None


def brand_from_iqvia_product(product: str) -> str:
    """'DILATREND (SIE)' → 'DILATREND'.
    'JARDIANCE DUO (B.I)' → 'JARDIANCE DUO' (puntos dentro de parens).
    Normaliza guiones a espacio: 'DIOVAN-D' → 'DIOVAN D'."""
    if not product:
        return ""
    # Acepta cualquier cosa entre paréntesis al final (incluyendo puntos)
    clean = re.sub(r"\s*\([^)]+\)\s*$", "", str(product)).strip()
    clean = clean.replace("-", " ")
    clean = re.sub(r"\s+", " ", clean)
    return clean


def _clean_num(s: str) -> str:
    """'10' → '10', '10.0' → '10', '3.125' → '3.125'."""
    try:
        f = float(s)
        if f.is_integer():
            return str(int(f))
        return str(f).rstrip("0").rstrip(".")
    except (ValueError, TypeError):
        return str(s)


def slug_dosis(mg: str | None) -> str:
    """Normaliza dosis para nombre de hoja."""
    if not mg:
        return "NA"
    if "/" in mg:
        parts = [_clean_num(p) for p in mg.split("/")]
        return f"{parts[0]}mg _ " + " _ ".join(f"{p}mg" for p in parts[1:])
    return f"{_clean_num(mg)}mg"


def sheet_name_for(marca: str, mg: str | None, max_len: int = 31) -> str:
    """Genera nombre de hoja tipo template. Excel limita a 31 chars."""
    name = f"{marca}_{slug_dosis(mg)}"
    return name[:max_len]


# ---------- Carga de datos ----------

def cargar_precios() -> pd.DataFrame:
    df = pd.read_excel(PRECIOS_XLSX, sheet_name=0)
    df.columns = [c.strip() for c in df.columns]
    # cast PVPs a float
    for col in (COL_PVP_ANT, COL_PVP_ACT):
        df[col] = pd.to_numeric(df[col], errors="coerce")
    df["Q Pres"] = pd.to_numeric(df["Q Pres"], errors="coerce").fillna(1)
    df["_lab_norm"] = df["Laboratorio"].map(norm_lab)
    df["_prod_norm"] = df["Producto"].map(norm_prod)
    df["_droga_norm"] = df["Droga"].map(norm_text)
    df["_mg"] = df["Presentacion"].map(extract_mg)
    df["_count"] = df["Presentacion"].map(extract_count)
    return df


def cargar_iqvia() -> pd.DataFrame:
    df = pd.read_excel(IQVIA_XLSX, sheet_name=0)
    df.columns = [c.replace("\n", " ").strip() for c in df.columns]
    df["Units MAT M 2026 March"] = pd.to_numeric(
        df["Units MAT M 2026 March"], errors="coerce"
    )
    df["_manuf_norm"] = df["Manufacturer"].map(norm_lab)
    df["_brand"] = df["Product"].map(brand_from_iqvia_product)
    df["_brand_norm"] = df["_brand"].map(norm_prod)
    df["_mg"] = df["Pack"].map(mg_from_iqvia_pack)
    df["_count"] = df["Pack"].map(count_from_iqvia_pack)
    df["_atc4"] = df["ATC IV"].astype(str).str.extract(r"^([A-Z]\d{2}[A-Z]?\d?)", expand=False)
    return df


def build_mat_lookup(iqvia: pd.DataFrame) -> tuple[dict, dict, set]:
    """Retorna:
    1) exact: dict (manuf_norm, brand_norm, mg_str, count) → MAT (con variantes de mg)
    2) by_brand: dict (manuf_norm, brand_norm) → list[(mg_original, count, MAT)]
    3) iqvia_labs: set con todos los lab_norm de IQVIA (para fuzzy match)
    """
    exact: dict = {}
    by_brand: dict = {}
    iqvia_labs: set = set()
    for _, r in iqvia.iterrows():
        u = r["Units MAT M 2026 March"]
        if pd.isna(u):
            continue
        lab = r["_manuf_norm"]
        br = r["_brand_norm"]
        mg = r["_mg"]
        ct = r["_count"]
        iqvia_labs.add(lab)
        for mg_v in mg_variants(mg):
            key = (lab, br, mg_v, ct)
            if key not in exact:
                exact[key] = float(u)
        by_brand.setdefault((lab, br), []).append((mg, ct, float(u)))
    return exact, by_brand, iqvia_labs


def mg_match_score(precios_mg: str | None, iqvia_mg: str | None,
                    precios_ct, iqvia_ct, tol: float = 0.6) -> int:
    """Puntaje de match entre dosis precios↔IQVIA. Mayor = mejor match.
    - 100: match componentwise exacto
    -  50: subset por primer componente (precios reporta solo el principal, ej ROVARTAL EZ '5')
    -  25: match por suma (precios reporta dosis total, ej ENTRESTO 50=24+26)
    -  +10: bonus si count coincide
    -   0: sin match de mg"""
    a = mg_numbers(precios_mg)
    b = mg_numbers(iqvia_mg)
    if not a or not b:
        return 0
    score = 0
    if len(a) == len(b) and all(abs(x - y) <= tol for x, y in zip(a, b)):
        score += 100
    elif len(a) < len(b) and all(abs(a[i] - b[i]) <= tol for i in range(len(a))):
        score += 50
    elif len(a) > len(b) and all(abs(a[i] - b[i]) <= tol for i in range(len(b))):
        score += 50
    elif len(a) != len(b) and abs(sum(a) - sum(b)) <= tol:
        score += 25
    if score > 0 and precios_ct == iqvia_ct:
        score += 10
    return score


def mg_numbers(mg_str: str | None) -> list[float]:
    """'160/12.5' → [160.0, 12.5]. '50' → [50.0]. None → []."""
    if not mg_str:
        return []
    try:
        return [float(p) for p in str(mg_str).split("/")]
    except ValueError:
        return []


def mg_compatible(siegfried_mg: str | None, otro_mg: str | None, tol: float = 0.6) -> bool:
    """True si 'otro_mg' es una presentación compatible con la dosis Siegfried.
    Casos: (a) igual exacto, (b) primer componente coincide y la presentación más corta es subset.
    Ej: Siegfried '5/10' ↔ competidor '5' → True (ROVARTAL EZ label la rosu solo).
        Siegfried '5/10' ↔ '10' → False (10 no es el primer componente)."""
    a = mg_numbers(siegfried_mg)
    b = mg_numbers(otro_mg)
    if not a or not b:
        return False
    n = min(len(a), len(b))
    return all(abs(a[i] - b[i]) <= tol for i in range(n))


def mg_matches_tol(precio_mg: str | None, iqvia_mg: str | None, tol: float = 0.6) -> bool:
    """True si dos cadenas de mg son equivalentes con tolerancia (para packs IQVIA truncados).
    - '12.5/1000' ≈ '1000/12' (orden y truncado) → True
    - '50' ≈ '24/26' (suma) → True
    - '5/850' ≈ '850/5' → True"""
    a = sorted(mg_numbers(precio_mg))
    b = sorted(mg_numbers(iqvia_mg))
    if not a or not b:
        return False
    if len(a) == len(b):
        return all(abs(x - y) <= tol for x, y in zip(a, b))
    # Asimetría: precios=suma, iqvia=componentes
    if len(a) == 1 and len(b) == 2:
        return abs(sum(b) - a[0]) <= tol
    if len(a) == 2 and len(b) == 1:
        return abs(sum(a) - b[0]) <= tol
    return False


def marca_base(producto: str) -> str:
    """Extrae la marca base de un producto Siegfried.
    'DILATREND AP' → 'DILATREND', 'EMPAX  MET' → 'EMPAX MET', 'ROXOLAN PLUS' → 'ROXOLAN PLUS'."""
    if not producto:
        return ""
    tokens = norm_text(producto).split()
    if not tokens:
        return ""
    first = tokens[0]
    # Sufijos que son parte de la marca (combinación)
    compuestas = {"MET", "PLUS"}
    if len(tokens) >= 2 and tokens[1] in compuestas:
        return f"{first} {tokens[1]}"
    return first


def es_producto_siegfried_cardio(producto: str) -> bool:
    """True si el producto Siegfried pertenece al portfolio cardio (primer token en whitelist)."""
    if not producto:
        return False
    tokens = norm_text(producto).split()
    return bool(tokens) and tokens[0] in CARDIO_BRANDS_SIEGFRIED


# ---------- Construcción de hojas ----------

def generar_hojas(
    precios: pd.DataFrame,
    iqvia: pd.DataFrame,
    mat_lookup: tuple[dict, dict, set],
) -> list[dict]:
    exact_lut, by_brand_lut, iqvia_labs = mat_lookup
    # Caché precios_lab → iqvia_lab para no re-fuzzy-matchear N veces
    lab_cache: dict = {}

    def _iqvia_lab_for(precios_lab_norm: str) -> str | None:
        if precios_lab_norm not in lab_cache:
            lab_cache[precios_lab_norm] = find_iqvia_lab(precios_lab_norm, iqvia_labs)
        return lab_cache[precios_lab_norm]
    """Una hoja por (marca_siegfried, dosis). Cada hoja: todos los competidores de esa
    (droga, mg) en el mercado, ordenados por MAT desc."""
    # Filas Siegfried en el portfolio cardio
    sieg = precios[
        (precios["_lab_norm"] == "SIEGFRIED")
        & (precios["Producto"].map(es_producto_siegfried_cardio))
    ].copy()
    sieg["_marca_hoja"] = sieg["Producto"].map(marca_base)

    grupos = []
    grouped = sieg.groupby(["_marca_hoja", "_mg", "_droga_norm"], dropna=False, sort=False)
    seen_keys: set[tuple[str, str]] = set()
    for (marca, dosis, droga), g in grouped:
        if not dosis or not droga:
            continue
        # Dedupe por (marca, dosis) — si hay dos drogas distintas en misma (marca, dosis)
        # tomo la mayor en unidades Siegfried (Q Pres) pero en la práctica es raro
        key = (marca, dosis)
        if key in seen_keys:
            continue
        seen_keys.add(key)

        # Competidores: misma Droga normalizada + dosis compatible
        # (acepta match exacto y subset por primer componente — ej ROVARTAL EZ '5' ↔ ROXOLAN PLUS '5/10')
        competidores = precios[
            (precios["_droga_norm"] == droga)
            & (precios["_mg"].apply(lambda m: mg_compatible(dosis, m)))
        ].copy()

        if competidores.empty:
            continue

        # Dedupe por (lab, producto, presentacion)
        competidores = competidores.drop_duplicates(
            subset=["_lab_norm", "_prod_norm", "Presentacion"]
        )

        def _mat_for(r):
            lab_iq = _iqvia_lab_for(r["_lab_norm"])
            if lab_iq is None:
                return None
            br = r["_prod_norm"]
            ct = r["_count"]
            mg = r["_mg"]

            # 1) Match exacto (lab, brand, mg variant, count)
            for mg_v in mg_variants(mg):
                k = (lab_iq, br, mg_v, ct)
                if k in exact_lut:
                    return exact_lut[k]
            # 2) Scored match dentro de (lab, brand) — elige la IQVIA row con mejor score
            candidates = by_brand_lut.get((lab_iq, br), [])
            best_mat = None
            best_score = 0
            for iq_mg, iq_ct, mat in candidates:
                s = mg_match_score(mg, iq_mg, ct, iq_ct)
                if s > best_score:
                    best_score = s
                    best_mat = mat
            # Umbral mínimo: 25 (match por sum o subset)
            if best_mat is not None and best_score >= 25:
                return best_mat
            return None

        # Consolidar múltiples Siegfried por bloque (quedarme con la de mayor MAT)
        sie_rows = competidores[competidores["_lab_norm"] == "SIEGFRIED"]
        if len(sie_rows) > 1:
            sie_with_mat = sie_rows.assign(
                _tmp_mat=sie_rows.apply(lambda r: _mat_for(r) or 0, axis=1)
            )
            keep_idx = sie_with_mat["_tmp_mat"].idxmax()
            drop_idx = [i for i in sie_rows.index if i != keep_idx]
            competidores = competidores.drop(index=drop_idx)

        competidores["_mat"] = competidores.apply(_mat_for, axis=1)

        # Orden: MAT desc (None al final), luego alfabético por lab
        competidores = competidores.sort_values(
            by=["_mat", "Laboratorio"],
            ascending=[False, True],
            na_position="last",
        ).reset_index(drop=True)

        grupos.append({
            "marca": marca,
            "dosis": dosis,
            "droga": droga,
            "competidores": competidores,
            "sheet_name": sheet_name_for(marca, dosis),
        })
    return grupos


# ---------- Escribir Excel ----------

SUBHDR_FILL = PatternFill("solid", fgColor="305496")  # azul oscuro más claro


def _dosis_display(dosis: str) -> str:
    """'160/12.5' → '160/12.5 mg'. '50' → '50 mg'."""
    if not dosis:
        return ""
    if "/" in dosis:
        return "/".join(_clean_num(p) for p in dosis.split("/")) + " mg"
    return f"{_clean_num(dosis)} mg"


def _escribir_bloque_dosis(ws, start_row: int, dosis: str, droga: str, comp: pd.DataFrame) -> int:
    """Escribe subheader + filas. Columnas: A=Lab, B=Producto, C=Dosis (nueva), D=Presentación,
    E=PVP Feb, F=PVP Abr, G=PVP/U, H=Var%, I=VS SIE, J=VS Líder, K=MAT, L=Share, M=Rank."""
    subhdr_row = start_row
    subhdr_label = f"Dosis: {_dosis_display(dosis)} — {droga.lower()}"
    ws.cell(row=subhdr_row, column=1, value=subhdr_label)
    ws.merge_cells(start_row=subhdr_row, start_column=1, end_row=subhdr_row, end_column=13)
    sub_cell = ws.cell(row=subhdr_row, column=1)
    sub_cell.fill = SUBHDR_FILL
    sub_cell.font = Font(bold=True, color="FFFFFF", size=11)
    sub_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[subhdr_row].height = 20

    first_row = subhdr_row + 1
    last_row = first_row + len(comp) - 1
    dosis_label = _dosis_display(dosis)

    sie_mask = comp["_lab_norm"] == "SIEGFRIED"
    sie_df_idx = comp.index[sie_mask].tolist()
    sie_excel_row = None
    if sie_df_idx:
        sie_excel_row = first_row + comp.index.get_loc(sie_df_idx[0])
    leader_excel_row = first_row

    for i, (_, r) in enumerate(comp.iterrows()):
        row = first_row + i
        is_sie = r["_lab_norm"] == "SIEGFRIED"
        is_leader = (i == 0)
        count = int(r["_count"]) if r["_count"] and not pd.isna(r["_count"]) else 1
        pvp_ant = float(r[COL_PVP_ANT]) if pd.notna(r[COL_PVP_ANT]) else None
        pvp_act = float(r[COL_PVP_ACT]) if pd.notna(r[COL_PVP_ACT]) else None
        mat_u = int(r["_mat"]) if pd.notna(r["_mat"]) else None

        lab_label = "Siegfried ★ SIE" if is_sie else str(r["Laboratorio"]).strip()

        ws.cell(row=row, column=1, value=lab_label)                      # A Lab
        ws.cell(row=row, column=2, value=str(r["Producto"]).strip())     # B Producto
        ws.cell(row=row, column=3, value=dosis_label)                    # C Dosis
        ws.cell(row=row, column=4, value=str(r["Presentacion"]).strip()) # D Presentación
        ws.cell(row=row, column=5, value=pvp_ant)                        # E PVP Feb
        ws.cell(row=row, column=6, value=pvp_act)                        # F PVP Abr
        ws.cell(row=row, column=7, value=f"=IFERROR(F{row}/{count},\"\")") # G PVP/U

        if is_sie:
            ws.cell(row=row, column=8, value=None)
        else:
            ws.cell(row=row, column=8, value=f"=IFERROR((F{row}-E{row})/E{row},\"\")")  # H Var%

        if is_sie or not sie_excel_row:
            ws.cell(row=row, column=9, value=None)
        else:
            ws.cell(row=row, column=9,
                    value=f"=IFERROR((G{row}-G{sie_excel_row})/G{sie_excel_row},\"\")")  # I VS SIE

        if is_leader:
            ws.cell(row=row, column=10, value="Líder")  # J VS Líder
        else:
            ws.cell(row=row, column=10,
                    value=f"=IFERROR((G{row}-G{leader_excel_row})/G{leader_excel_row},\"\")")

        ws.cell(row=row, column=11, value=mat_u)  # K MAT
        ws.cell(row=row, column=12,
                value=f"=IFERROR(K{row}/SUM($K${first_row}:$K${last_row}),\"\")")  # L Share
        ws.cell(row=row, column=13, value=i + 1)  # M Ranking

        # Formatos
        for c in (5, 6, 7):
            ws.cell(row=row, column=c).number_format = "#,##0.00"
        ws.cell(row=row, column=11).number_format = "#,##0"
        ws.cell(row=row, column=13).number_format = "0"
        for c in (8, 9, 12):
            ws.cell(row=row, column=c).number_format = "0.0%"
        ws.cell(row=row, column=10).number_format = "0.0%" if not is_leader else "General"

        if is_sie:
            for c in range(1, 14):
                ws.cell(row=row, column=c).fill = SIE_FILL
                ws.cell(row=row, column=c).font = SIE_FONT
        if is_leader:
            ws.cell(row=row, column=10).fill = LEADER_FILL
            ws.cell(row=row, column=10).font = Font(bold=True)

    return last_row + 2


def escribir_excel(grupos: list[dict], precios: pd.DataFrame, out_path: Path):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # _Resumen
    ws_resumen = wb.create_sheet("_Resumen")
    ws_resumen.append([
        "Marca Siegfried", "Dosis", "Droga", "Hoja", "# Competidores",
        "PVP Siegfried Abr", "Líder por MAT", "MAT Líder",
    ])
    for c in range(1, 9):
        ws_resumen.cell(row=1, column=c).fill = HDR_FILL
        ws_resumen.cell(row=1, column=c).font = HDR_FONT

    headers = [
        "Laboratorio", "Producto", "Dosis", "Presentación",
        LABEL_PVP_ANT, LABEL_PVP_ACT, "PVP/Unidad",
        "Var %", "VS SIE", "VS Líder",
        "IQVIA MAT U.", "Share %", "Ranking",
    ]

    # Agrupar por marca
    def dosis_key(dosis: str) -> float:
        try:
            return float(dosis.split("/")[0])
        except (ValueError, AttributeError):
            return 0.0

    from collections import defaultdict
    por_marca: dict[str, list[dict]] = defaultdict(list)
    for g in grupos:
        por_marca[g["marca"]].append(g)
    for marca in por_marca:
        por_marca[marca].sort(key=lambda g: dosis_key(g["dosis"]))

    # Orden alfabético de marcas
    for marca in sorted(por_marca.keys()):
        grupos_marca = por_marca[marca]

        # Nombre de hoja ≤ 31 chars
        sheet_name = marca[:31]
        if sheet_name in wb.sheetnames:
            sheet_name = f"{marca[:29]}_2"
        ws = wb.create_sheet(sheet_name)

        # Header (fila 1)
        for col_idx, h in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=h)
            cell.fill = HDR_FILL
            cell.font = HDR_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # Anchos
        for letter, w in COL_WIDTHS.items():
            ws.column_dimensions[letter].width = w
        ws.freeze_panes = "A2"

        # Escribir cada bloque de dosis
        next_row = 2
        for grupo in grupos_marca:
            next_row = _escribir_bloque_dosis(
                ws, next_row, grupo["dosis"], grupo["droga"], grupo["competidores"]
            )

            # Fila en _Resumen
            comp = grupo["competidores"]
            sie_mask = comp["_lab_norm"] == "SIEGFRIED"
            sie_idx = comp.index[sie_mask].tolist()
            sie_pvp_act = comp.loc[sie_idx[0], COL_PVP_ACT] if sie_idx else None
            lider_row = comp.iloc[0]
            ws_resumen.append([
                marca,
                grupo["dosis"],
                grupo["droga"].lower() if grupo["droga"] else "",
                sheet_name,
                len(comp),
                float(sie_pvp_act) if sie_pvp_act is not None and pd.notna(sie_pvp_act) else None,
                f"{lider_row['Laboratorio']} — {lider_row['Producto']}",
                int(lider_row["_mat"]) if pd.notna(lider_row["_mat"]) else None,
            ])

    # Formato resumen
    for r in range(2, ws_resumen.max_row + 1):
        ws_resumen.cell(row=r, column=6).number_format = "#,##0.00"
        ws_resumen.cell(row=r, column=8).number_format = "#,##0"
    for i, w in enumerate([22, 18, 32, 22, 15, 18, 45, 15], start=1):
        ws_resumen.column_dimensions[get_column_letter(i)].width = w
    ws_resumen.freeze_panes = "A2"

    # _Resumen al principio (ya está primero porque se creó primero)
    wb.save(out_path)


# ---------- Main ----------

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--dry-run", action="store_true", help="No escribe el Excel, solo imprime resumen")
    ap.add_argument("--output", default=str(OUT_XLSX), help="Ruta del Excel de salida")
    args = ap.parse_args()
    out_path = Path(args.output)

    print("1) Cargando precios…")
    precios = cargar_precios()
    print(f"   {len(precios)} filas, {precios['Laboratorio'].nunique()} laboratorios")

    print("2) Cargando IQVIA…")
    iqvia = cargar_iqvia()
    print(f"   {len(iqvia)} filas, {iqvia['ATC IV'].nunique()} códigos ATC")

    print("3) Construyendo lookup MAT…")
    mat_lookup = build_mat_lookup(iqvia)
    print(f"   {len(mat_lookup)} claves MAT")

    print(f"4) Filtrando productos Siegfried cardio (whitelist: {len(CARDIO_BRANDS_SIEGFRIED)} marcas)…")
    n_sieg_cardio = precios[
        (precios["_lab_norm"] == "SIEGFRIED")
        & (precios["Producto"].map(es_producto_siegfried_cardio))
    ].shape[0]
    print(f"   {n_sieg_cardio} presentaciones Siegfried cardio en precios")

    print("5) Generando hojas…")
    grupos = generar_hojas(precios, iqvia, mat_lookup)
    print(f"   {len(grupos)} hojas (una por marca+dosis)")
    for g in grupos:
        n = len(g["competidores"])
        sie_present = (g["competidores"]["_lab_norm"] == "SIEGFRIED").any()
        lider = g["competidores"].iloc[0]
        print(f"     {g['sheet_name']:35s}  droga={g['droga'][:25]:25s}  n={n:3d}  líder={lider['Laboratorio']}—{lider['Producto']}")

    if args.dry_run:
        print("\n(dry-run) Sin escribir Excel.")
        return

    print(f"\n6) Escribiendo {out_path.name}…")
    escribir_excel(grupos, precios, out_path)
    print(f"   ✓ guardado: {out_path}")


if __name__ == "__main__":
    main()
